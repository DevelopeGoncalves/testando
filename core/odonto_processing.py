"""
Processamento das planilhas ODONTO (Odontoprev), no mesmo padrão do ANBIMA.

Recebe os dois relatórios exportados do site da Odontoprev — "Corretora Total
Vidas PF" e "Corretora Total Vidas PME" — cruza com o cadastro de Colaboradores
e Unidades (Agência/CID) do banco e devolve o arquivo "PRODUÇÃO <MÊS> <ANO>.xlsx"
já montado em memória, com as abas "Pessoa Física" e "PJ".

As colunas aproveitadas de cada relatório são as marcadas em amarelo pelo
usuário nas planilhas de origem:
  PF : DT_VENDA, NOME_FORCA, CPF_FORCA, TIPO_PLANO, VALOR_VENDA, STATUS_PROPOSTA
  PJ : DT_VENDA, CPF, VENDEDOR, VALOR_VENDA, STATUS

Os campos "Indicador", "Matrícula", "CID" e "Agência" substituem os PROCVs da
planilha manual: são resolvidos a partir de Base > Formulários > Colaboradores
(nome e matrícula) e do cadastro de Unidades / Agência CID (CID e nome da
agência), tendo o CPF da força de vendas como chave de busca.
"""
import re
import unicodedata
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Status considerado "produção válida" conforme a aba Orientações da planilha.
STATUS_CONCLUIDO = 'PROPOSTA CONCLUIDA COM SUCESSO'

# Valores de produção por tipo de plano PF (replica a fórmula da planilha:
# =IF(TIPO="MENSAL";54,99;IF(TIPO="ANUAL";659,88;0))).
VALOR_MENSAL_PF = 54.99
VALOR_ANUAL_PF = 659.88

MESES_ABREV = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
               'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

# Colunas mínimas que cada relatório precisa ter para o processamento rodar.
COLUNAS_PF = ['DT_VENDA', 'NOME_FORCA', 'CPF_FORCA', 'TIPO_PLANO', 'STATUS_PROPOSTA']
COLUNAS_PJ = ['NUM_PROPOSTA', 'DT_VENDA', 'CPF', 'VENDEDOR', 'VALOR_VENDA', 'STATUS']

CABECALHO_PF = ['CPF_FORCA', 'Indicador', 'Matrícula', 'CID', 'Realizado',
                'Contagem de matriculas', 'DT_VENDA', 'TIPO', 'Agência']
CABECALHO_PJ = ['CPF_FORCA', 'Indicador', 'Matrícula', 'CID', 'Realizado', 'Agência']

FORMATO_MOEDA = 'R$ #,##0.00'


# --------------------------------------------------------------------------- #
# Normalizações
# --------------------------------------------------------------------------- #
def _sem_acento(texto):
    """Maiúsculas, sem acento e com espaços colapsados — usado nas comparações."""
    if texto is None:
        return ''
    texto = unicodedata.normalize('NFKD', str(texto)).encode('ascii', 'ignore').decode()
    return ' '.join(texto.upper().split())


def _so_digitos(valor):
    if valor is None:
        return ''
    return re.sub(r'\D', '', str(valor))


def _cpf_normalizado(valor):
    """CPF apenas com dígitos e completado com zeros à esquerda (11 posições)."""
    digitos = _so_digitos(valor)
    if not digitos:
        return ''
    # A exportação da Odontoprev às vezes perde o zero inicial do CPF.
    return digitos.zfill(11) if len(digitos) <= 11 else digitos


def _para_numero(valor):
    """Converte o valor da venda vindo como texto ('659.88', '1.234,56') em float."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return float(valor)

    texto = str(valor).strip().replace('R$', '').replace(' ', '')
    if not texto:
        return 0.0

    if ',' in texto and '.' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    elif ',' in texto:
        texto = texto.replace(',', '.')

    try:
        return float(texto)
    except ValueError:
        return 0.0


def _matricula_exibicao(matricula):
    """A planilha mostra a matrícula sem os zeros à esquerda (030085411 -> 30085411)."""
    digitos = _so_digitos(matricula).lstrip('0')
    return int(digitos) if digitos else None


# --------------------------------------------------------------------------- #
# Leitura dos relatórios da Odontoprev
# --------------------------------------------------------------------------- #
def ler_relatorio_odonto(arquivo):
    """
    Lê o .xls exportado pela Odontoprev.

    O arquivo traz duas linhas de identificação da corretora antes do cabeçalho
    real, então a linha de cabeçalho é localizada pelo conteúdo em vez de ser
    fixada em uma posição.
    """
    bruto = None
    for engine in ('xlrd', 'openpyxl'):
        try:
            arquivo.seek(0)
            bruto = pd.read_excel(arquivo, engine=engine, header=None, dtype=str)
            break
        except ImportError:
            continue
        except Exception:
            continue

    if bruto is None:
        raise ValueError(
            'Não foi possível ler o arquivo enviado. Verifique se é o relatório '
            'da Odontoprev em formato Excel (.xls ou .xlsx) e exporte novamente.'
        )

    linha_cabecalho = None
    for indice, linha in bruto.iterrows():
        valores = {_sem_acento(v) for v in linha.tolist()}
        if 'NUM_PROPOSTA' in valores and 'DT_VENDA' in valores:
            linha_cabecalho = indice
            break

    if linha_cabecalho is None:
        raise ValueError(
            'Cabeçalho não localizado na planilha (não foi encontrada a coluna '
            'NUM_PROPOSTA). Envie o relatório exportado pelo site da Odontoprev.'
        )

    colunas = [_sem_acento(c) for c in bruto.iloc[linha_cabecalho].tolist()]
    df = bruto.iloc[linha_cabecalho + 1:].copy()
    df.columns = colunas
    df = df.loc[:, [c for c in df.columns if c]]
    df = df.dropna(how='all').reset_index(drop=True)
    return df


# --------------------------------------------------------------------------- #
# Cruzamento com a base (Colaboradores + Agência CID)
# --------------------------------------------------------------------------- #
class BaseColaboradores:
    """
    Índice de consulta em memória sobre o cadastro de Colaboradores.

    Substitui os PROCVs da planilha manual: procura primeiro pelo CPF da força
    de vendas e, quando o CPF não está preenchido na base, cai para o nome —
    inclusive quando a Odontoprev trunca o nome em 30 caracteres.
    """

    def __init__(self, colaboradores):
        self.por_cpf = {}
        self.por_nome = {}
        self.nomes = []

        for colaborador in colaboradores:
            cpf = _cpf_normalizado(colaborador.cpf)
            if cpf:
                self._registrar(self.por_cpf, cpf, colaborador)

            for nome in (colaborador.colaborador, colaborador.nome_social):
                nome_chave = _sem_acento(nome)
                if nome_chave:
                    self._registrar(self.por_nome, nome_chave, colaborador)
                    self.nomes.append((nome_chave, colaborador))

    @staticmethod
    def _registrar(indice, chave, colaborador):
        atual = indice.get(chave)
        # Havendo repetição, mantém o colaborador ativo (o inativo é o histórico).
        if atual is None or (atual.inativo and not colaborador.inativo):
            indice[chave] = colaborador

    def buscar(self, cpf, nome):
        colaborador = self.por_cpf.get(_cpf_normalizado(cpf))
        if colaborador:
            return colaborador

        nome_chave = _sem_acento(nome)
        if not nome_chave:
            return None

        colaborador = self.por_nome.get(nome_chave)
        if colaborador:
            return colaborador

        # O relatório PF corta o nome em 30 caracteres ("Miguel Rogerio Souza
        # Strecht J"), então tenta casar por prefixo quando não há ambiguidade.
        if len(nome_chave) >= 15:
            candidatos = {c.pk: c for chave, c in self.nomes if chave.startswith(nome_chave)}
            if len(candidatos) == 1:
                return next(iter(candidatos.values()))

        return None


def _dados_do_colaborador(colaborador, nome_planilha):
    """Monta as colunas Indicador / Matrícula / CID / Agência de uma linha."""
    if colaborador is None:
        nome = str(nome_planilha or '').strip()
        return {'Indicador': nome or '-', 'Matrícula': None, 'CID': None, 'Agência': None}

    unidade = colaborador.unidade
    return {
        'Indicador': colaborador.nome_social or colaborador.colaborador,
        'Matrícula': _matricula_exibicao(colaborador.matricula),
        'CID': _matricula_exibicao(unidade.cid_unidade) if unidade else None,
        'Agência': unidade.unidade if unidade else None,
    }


# --------------------------------------------------------------------------- #
# Montagem das abas
# --------------------------------------------------------------------------- #
def _filtrar(df, coluna_status, mes, ano, apenas_concluidas):
    """Aplica o filtro de status e mantém apenas as vendas do mês de referência."""
    df = df.copy()

    if apenas_concluidas:
        df = df[df[coluna_status].apply(lambda s: _sem_acento(s) == STATUS_CONCLUIDO)]

    df['_data_venda'] = pd.to_datetime(df['DT_VENDA'], dayfirst=True, errors='coerce')
    df = df[df['_data_venda'].notna()]
    df = df[(df['_data_venda'].dt.month == mes) & (df['_data_venda'].dt.year == ano)]
    return df


def _montar_pf(df, base, mes, ano, apenas_concluidas):
    df = _filtrar(df, 'STATUS_PROPOSTA', mes, ano, apenas_concluidas)

    linhas = []
    nao_encontrados = []
    for _, registro in df.iterrows():
        cpf = _cpf_normalizado(registro.get('CPF_FORCA'))
        nome_planilha = registro.get('NOME_FORCA')
        colaborador = base.buscar(cpf, nome_planilha)
        if colaborador is None:
            nao_encontrados.append((cpf, str(nome_planilha or '').strip(), 'PF'))

        tipo = _sem_acento(registro.get('TIPO_PLANO'))
        if tipo == 'MENSAL':
            realizado = VALOR_MENSAL_PF
        elif tipo == 'ANUAL':
            realizado = VALOR_ANUAL_PF
        else:
            # Tipo fora do padrão: usa o valor da própria venda em vez de zerar.
            realizado = _para_numero(registro.get('VALOR_VENDA'))

        linha = {'CPF_FORCA': cpf}
        linha.update(_dados_do_colaborador(colaborador, nome_planilha))
        linha['Realizado'] = realizado
        linha['DT_VENDA'] = registro['_data_venda'].strftime('%d/%m/%Y')
        linha['TIPO'] = str(registro.get('TIPO_PLANO') or '').strip()
        linhas.append(linha)

    resultado = pd.DataFrame(linhas, columns=CABECALHO_PF)
    if resultado.empty:
        return resultado, nao_encontrados

    # "Contagem de matriculas" reproduz o COUNTIF(C:C;C2) da planilha manual.
    # Fica em branco nas linhas sem matrícula (força de vendas fora da base).
    contagem = resultado['Matrícula'].map(resultado['Matrícula'].value_counts())
    resultado['Contagem de matriculas'] = contagem.astype('Int64')

    resultado['_ordem'] = resultado['Indicador'].map(_sem_acento)
    resultado = resultado.sort_values(by=['_ordem', 'DT_VENDA']).drop(columns='_ordem')
    return resultado.reset_index(drop=True), nao_encontrados


def _montar_pj(df, base, mes, ano, apenas_concluidas):
    df = _filtrar(df, 'STATUS', mes, ano, apenas_concluidas)

    # Uma proposta PME aparece repetida uma vez por vida contratada, mas o
    # VALOR_VENDA já é o total — então só a primeira linha de cada proposta entra.
    df = df.drop_duplicates(subset=['NUM_PROPOSTA'], keep='first')

    linhas = []
    nao_encontrados = []
    for _, registro in df.iterrows():
        cpf = _cpf_normalizado(registro.get('CPF'))
        nome_planilha = registro.get('VENDEDOR')
        colaborador = base.buscar(cpf, nome_planilha)
        if colaborador is None:
            nao_encontrados.append((cpf, str(nome_planilha or '').strip(), 'PJ'))

        linha = {'CPF_FORCA': cpf}
        linha.update(_dados_do_colaborador(colaborador, nome_planilha))
        linha['Realizado'] = _para_numero(registro.get('VALOR_VENDA'))
        linhas.append(linha)

    resultado = pd.DataFrame(linhas, columns=CABECALHO_PJ)
    if resultado.empty:
        return resultado, nao_encontrados

    resultado['_ordem'] = resultado['Indicador'].map(_sem_acento)
    resultado = resultado.sort_values(by='_ordem').drop(columns='_ordem')
    return resultado.reset_index(drop=True), nao_encontrados


# --------------------------------------------------------------------------- #
# Formatação da saída
# --------------------------------------------------------------------------- #
def _formatar_aba(sheet, colunas):
    preenchimento = PatternFill('solid', fgColor='6F42C1')
    for indice, _ in enumerate(colunas, start=1):
        celula = sheet.cell(row=1, column=indice)
        celula.font = Font(bold=True, color='FFFFFF')
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal='center', vertical='center')

    if 'Realizado' in colunas:
        coluna_realizado = colunas.index('Realizado') + 1
        for linha in range(2, sheet.max_row + 1):
            celula = sheet.cell(row=linha, column=coluna_realizado)
            if isinstance(celula.value, (int, float)):
                celula.number_format = FORMATO_MOEDA

    for indice, titulo in enumerate(colunas, start=1):
        largura = len(str(titulo))
        for linha in range(2, sheet.max_row + 1):
            valor = sheet.cell(row=linha, column=indice).value
            if valor is not None:
                largura = max(largura, len(str(valor)))
        sheet.column_dimensions[get_column_letter(indice)].width = min(largura + 4, 60)

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions


# --------------------------------------------------------------------------- #
# Entrada principal
# --------------------------------------------------------------------------- #
def processar_planilhas_odonto(arquivo_pf, arquivo_pj, mes, ano, apenas_concluidas, colaboradores):
    """
    arquivo_pf / arquivo_pj: arquivos enviados (relatórios PF e PME da Odontoprev).
    mes / ano: mês e ano de referência da produção.
    apenas_concluidas: quando True, considera só "Proposta concluida com sucesso".
    colaboradores: queryset de Colaborador (usar select_related('unidade')).

    Retorna um dict:
      {'ok': True, 'buffer': BytesIO, 'nome_arquivo': str, 'total_pf': int,
       'total_pj': int, 'total_geral': float, 'nao_encontrados': [dict, ...]}
      {'ok': False, 'colunas_faltantes_pf': [...], 'colunas_faltantes_pj': [...]}
    """
    df_pf = ler_relatorio_odonto(arquivo_pf)
    df_pj = ler_relatorio_odonto(arquivo_pj)

    faltantes_pf = [c for c in COLUNAS_PF if c not in df_pf.columns]
    faltantes_pj = [c for c in COLUNAS_PJ if c not in df_pj.columns]
    if faltantes_pf or faltantes_pj:
        return {'ok': False, 'colunas_faltantes_pf': faltantes_pf, 'colunas_faltantes_pj': faltantes_pj}

    base = BaseColaboradores(colaboradores)

    resultado_pf, nao_encontrados_pf = _montar_pf(df_pf, base, mes, ano, apenas_concluidas)
    resultado_pj, nao_encontrados_pj = _montar_pj(df_pj, base, mes, ano, apenas_concluidas)

    # Lista única de forças de venda que não têm correspondência na base.
    nao_encontrados = []
    vistos = set()
    for cpf, nome, origem in nao_encontrados_pf + nao_encontrados_pj:
        if (cpf, nome) in vistos:
            continue
        vistos.add((cpf, nome))
        nao_encontrados.append({'cpf': cpf, 'nome': nome, 'origem': origem})
    nao_encontrados.sort(key=lambda item: (item['origem'], _sem_acento(item['nome'])))

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        resultado_pf.to_excel(writer, sheet_name='Pessoa Física', index=False)
        resultado_pj.to_excel(writer, sheet_name='PJ', index=False)
        _formatar_aba(writer.sheets['Pessoa Física'], CABECALHO_PF)
        _formatar_aba(writer.sheets['PJ'], CABECALHO_PJ)

        # As linhas sem correspondência saem com Matrícula/CID/Agência em branco;
        # esta aba lista quem precisa ser acertado no cadastro de Colaboradores.
        if nao_encontrados:
            df_pendencias = pd.DataFrame(nao_encontrados).rename(
                columns={'cpf': 'CPF_FORCA', 'nome': 'Nome na planilha', 'origem': 'Origem'}
            )[['CPF_FORCA', 'Nome na planilha', 'Origem']]
            df_pendencias.to_excel(writer, sheet_name='Pendências', index=False)
            _formatar_aba(writer.sheets['Pendências'], list(df_pendencias.columns))
    buffer.seek(0)

    total_geral = float(resultado_pf['Realizado'].sum() + resultado_pj['Realizado'].sum())

    return {
        'ok': True,
        'buffer': buffer,
        'nome_arquivo': f'PRODUÇÃO ODONTO {MESES_ABREV[mes - 1]} {ano}.xlsx',
        'total_pf': int(len(resultado_pf)),
        'total_pj': int(len(resultado_pj)),
        'total_geral': round(total_geral, 2),
        'nao_encontrados': nao_encontrados,
    }
