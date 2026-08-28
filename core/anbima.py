import math
from io import BytesIO

import pandas as pd
from django.utils import timezone
from openpyxl.styles import Font

COLUNAS_PARA_EXTRACAO = ['CPF', 'Fundo', 'Reserva', 'Data de contratacao', 'Cidade', 'UF']

# Fundos que mudaram de nome/gestora ao longo do tempo mas representam o mesmo produto.
FUNDOS_IGUAIS = {
    'CAPITÂNIA CREDPREVIDÊNCIA ICATU FIC DE FIRF CP': 'CAPITÂNIA CREDPREVIDÊNCIA ICATU FIC DE FIF RF CP RESP LTDA',
    'SCHRODER ICATU PREVIDÊNCIA LOW VOL FIF MULTIMERCADO RESP LTD': 'RIZA ICATU PREVIDÊNCIA LOW VOL FIF MULTIMERCADO RESP LTDA',
    'SCHRODER ICATU PREV LOW VOL FIM': 'RIZA ICATU PREVIDÊNCIA LOW VOL FIF MULTIMERCADO RESP LTDA',
}


def _truncar_6_casas(valor):
    """Trunca (corta, sem arredondar) o valor em 6 casas decimais."""
    if valor is None:
        return valor
    return math.trunc(valor * 1_000_000) / 1_000_000


def _ajustar_largura_colunas(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        sheet.column_dimensions[column].width = min(max_length + 2, 100)


def _formatar_negrito_linha(sheet, linha):
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(linha, col_idx)
        if cell.value is not None:
            cell.font = Font(bold=True)


def _formatar_colunas_numero(sheet, coluna_volume_nome, coluna_quantidade_nome, coluna_quantidade_contas_nome):
    col_indices = {}
    for col_idx, cell in enumerate(sheet[1]):
        col_name = cell.value
        if col_name == coluna_volume_nome:
            col_indices['volume'] = col_idx + 1
        elif col_name == coluna_quantidade_nome:
            col_indices['quantidade_unica'] = col_idx + 1
        elif col_name == coluna_quantidade_contas_nome:
            col_indices['quantidade_contas'] = col_idx + 1

    # Remove o registro inteiro quando o Volume trunca para zero.
    linhas_para_remover = []
    if 'volume' in col_indices:
        for row_idx in range(2, sheet.max_row + 1):
            valor_volume = sheet.cell(row=row_idx, column=col_indices['volume']).value
            if isinstance(valor_volume, (int, float)) and _truncar_6_casas(valor_volume) == 0:
                linhas_para_remover.append(row_idx)

    for row_idx in reversed(linhas_para_remover):
        sheet.delete_rows(row_idx, 1)

    for row_idx in range(2, sheet.max_row + 1):
        if 'volume' in col_indices:
            cell_volume = sheet.cell(row=row_idx, column=col_indices['volume'])
            if isinstance(cell_volume.value, (int, float)):
                cell_volume.number_format = '#,##0.000000'
        if 'quantidade_unica' in col_indices:
            cell_quantidade = sheet.cell(row=row_idx, column=col_indices['quantidade_unica'])
            if isinstance(cell_quantidade.value, (int, float)):
                cell_quantidade.number_format = '#,##0'
        if 'quantidade_contas' in col_indices:
            cell_quantidade_contas = sheet.cell(row=row_idx, column=col_indices['quantidade_contas'])
            if isinstance(cell_quantidade_contas.value, (int, float)):
                cell_quantidade_contas.number_format = '#,##0'


def _mes_ano_um_mes_antes(ano, mes):
    mes -= 1
    if mes == 0:
        mes = 12
        ano -= 1
    return ano, mes


def processar_planilha_anbima(df, data_limite, fundos_cadastrados, estados_cadastrados):
    
    fundos_cadastrados = list(fundos_cadastrados)
    estados_cadastrados = list(estados_cadastrados)

    colunas_faltantes = [c for c in COLUNAS_PARA_EXTRACAO if c not in df.columns]
    if colunas_faltantes:
        return {'ok': False, 'colunas_faltantes': colunas_faltantes}

    df_extraido = df[COLUNAS_PARA_EXTRACAO].copy()
    df_extraido['Fundo'] = df_extraido['Fundo'].replace(FUNDOS_IGUAIS)

    df_extraido['Data de contratacao'] = pd.to_datetime(
        df_extraido['Data de contratacao'], dayfirst=True, errors='coerce'
    ).dt.date

    df_extraido['Reserva'] = pd.to_numeric(
        df_extraido['Reserva'].astype(str)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
        .str.replace(' ', '', regex=False),
        errors='coerce'
    ).fillna(0)

    df_extraido['CPF'] = pd.to_numeric(df_extraido['CPF'], errors='coerce').fillna(0).astype('Int64')

    df_filtrado = df_extraido.copy()
    if data_limite:
        df_filtrado = df_filtrado[df_filtrado['Data de contratacao'].apply(lambda x: x is not None and x < data_limite)]
    else:
        agora = timezone.localtime(timezone.now())
        df_filtrado = df_filtrado[df_filtrado['Data de contratacao'].apply(
            lambda x: x is not None and (x.month != agora.month or x.year != agora.year)
        )]

    ordem_fundos = [f.nome_fundo for f in fundos_cadastrados]
    fundos_conhecidos = set(ordem_fundos)
    fundos_na_planilha = df_filtrado['Fundo'].dropna().unique()
    fundos_desconhecidos = sorted(f for f in fundos_na_planilha if f not in fundos_conhecidos)
    if fundos_desconhecidos:
        return {'ok': False, 'fundos_desconhecidos': fundos_desconhecidos}

    # --- TABELA AGREGADA POR UF ---
    dados_agregados = []
    indice_sp = None
    for i, estado in enumerate(estados_cadastrados):
        df_uf = df_filtrado[df_filtrado['UF'] == estado.uf]
        dados_agregados.append({
            'LOCAL': estado.uf_estado,
            'Soma_de_Reserva': df_uf['Reserva'].sum().round(2),
            'Contagem_de_CPF': df_uf['CPF'].nunique(),
            'Quantidade_por_contas': df_uf['CPF'].count(),
        })
        if estado.uf == 'SP':
            indice_sp = i

    total_soma_reserva = sum(item['Soma_de_Reserva'] for item in dados_agregados)
    total_contagem_cpf = df_filtrado['CPF'].nunique()
    total_contagem_contas = df_filtrado['CPF'].count()
    df_total_uf = pd.DataFrame([{
        'LOCAL': 'Total Previdência Aberta',
        'Soma_de_Reserva': total_soma_reserva,
        'Contagem_de_CPF': total_contagem_cpf,
        'Quantidade_por_contas': total_contagem_contas,
    }])

    df_agregado_estados = pd.DataFrame(dados_agregados)

    # Divide SP em Região Metropolitana / Interior, logo após a linha de SP.
    df_sp = df_filtrado[df_filtrado['UF'] == 'SP'].copy()
    df_sp_capital = df_sp[df_sp['Cidade'].astype(str).str.lower().str.strip().isin(['são paulo', 'sao paulo'])].copy()
    df_sp_interior = df_sp[~df_sp['Cidade'].astype(str).str.lower().str.strip().isin(['são paulo', 'sao paulo'])].copy()

    df_sp_resultados = pd.DataFrame([
        {
            'LOCAL': '    SP - Região Metropolitana',
            'Soma_de_Reserva': df_sp_capital['Reserva'].sum().round(2),
            'Contagem_de_CPF': df_sp_capital['CPF'].nunique(),
            'Quantidade_por_contas': df_sp_capital['CPF'].count(),
        },
        {
            'LOCAL': '    SP - São Paulo Interior',
            'Soma_de_Reserva': df_sp_interior['Reserva'].sum().round(2),
            'Contagem_de_CPF': df_sp_interior['CPF'].nunique(),
            'Quantidade_por_contas': df_sp_interior['CPF'].count(),
        },
    ])

    if indice_sp is not None:
        df_agregado_final_uf = pd.concat([
            df_agregado_estados.iloc[:indice_sp + 1],
            df_sp_resultados,
            df_agregado_estados.iloc[indice_sp + 1:],
        ], ignore_index=True)
    else:
        df_agregado_final_uf = pd.concat([df_agregado_estados, df_sp_resultados], ignore_index=True)

    df_agregado_uf = pd.concat([df_total_uf, df_agregado_final_uf], ignore_index=True)
    df_agregado_uf.rename(columns={
        'LOCAL': 'Estado (UF)',
        'Soma_de_Reserva': 'Volume em R$/mil',
        'Contagem_de_CPF': 'Quantidade',
        'Quantidade_por_contas': 'quantidade por contas',
    }, inplace=True)
    df_agregado_uf['Volume em R$/mil'] = (df_agregado_uf['Volume em R$/mil'] / 1000).apply(_truncar_6_casas)

    # --- TABELA POR FUNDO ---
    df_valores_por_fundo = df_filtrado.groupby('Fundo', as_index=False).agg(
        Soma_de_Reserva=('Reserva', 'sum'),
        Contagem_de_CPF=('CPF', 'nunique'),
        Quantidade_por_contas=('CPF', 'count'),
    )

    df_ordem_fundos = pd.DataFrame({'Fundo': ordem_fundos})
    df_valores_ordenados = pd.merge(df_ordem_fundos, df_valores_por_fundo, on='Fundo', how='left').fillna(
        {'Soma_de_Reserva': 0, 'Contagem_de_CPF': 0, 'Quantidade_por_contas': 0}
    )
    df_valores_ordenados['Contagem_de_CPF'] = df_valores_ordenados['Contagem_de_CPF'].astype(int)
    df_valores_ordenados['Quantidade_por_contas'] = df_valores_ordenados['Quantidade_por_contas'].astype(int)

    total_reserva_fundo = df_filtrado['Reserva'].sum().round(2)
    df_total_fundo = pd.DataFrame([{
        'Fundo': 'Total Previdência Aberta',
        'Soma_de_Reserva': total_reserva_fundo,
        'Contagem_de_CPF': df_filtrado['CPF'].nunique(),
        'Quantidade_por_contas': df_filtrado['CPF'].count(),
    }])

    codigos_anbima_dict = {f.nome_fundo: (f.codigo_anbima or '') for f in fundos_cadastrados}
    cnpj_dict = {f.nome_fundo: (f.cnpj_fundo or '') for f in fundos_cadastrados}

    df_valores_ordenados['CÓDIGO ANBIMA'] = df_valores_ordenados['Fundo'].map(codigos_anbima_dict)
    df_valores_ordenados['CNPJ Fundo'] = df_valores_ordenados['Fundo'].map(cnpj_dict)

    df_total_fundo['CÓDIGO ANBIMA'] = ''
    df_total_fundo['CNPJ Fundo'] = ''

    df_valores_por_fundo_final = pd.concat([df_total_fundo, df_valores_ordenados], ignore_index=True)
    df_valores_por_fundo_final['Volume R$/mil'] = (df_valores_por_fundo_final['Soma_de_Reserva'] / 1000).apply(_truncar_6_casas)

    df_valores_por_fundo_final = df_valores_por_fundo_final[
        ['CNPJ Fundo', 'Fundo', 'CÓDIGO ANBIMA', 'Volume R$/mil', 'Contagem_de_CPF', 'Quantidade_por_contas']
    ]
    df_valores_por_fundo_final.rename(columns={
        'Fundo': 'Nome do Fundo',
        'Contagem_de_CPF': 'Quantidade CPFs únicos',
        'Quantidade_por_contas': 'Quantidade por Contas',
    }, inplace=True)

    # --- NOME DO ARQUIVO (mês/ano anterior à data de corte, ou ao mês atual) ---
    if data_limite:
        ano_nome, mes_nome = _mes_ano_um_mes_antes(data_limite.year, data_limite.month)
    else:
        agora = timezone.localtime(timezone.now())
        ano_nome, mes_nome = _mes_ano_um_mes_antes(agora.year, agora.month)
    nome_arquivo = f"ANBIMA_{mes_nome:02d}_{ano_nome}.xlsx"

    # --- GERA O ARQUIVO EM MEMÓRIA ---
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_agregado_uf.to_excel(writer, sheet_name='Por Estado', index=False)
        df_valores_por_fundo_final.to_excel(writer, sheet_name='Valores por Fundos', index=False)

        worksheet_estados = writer.sheets['Por Estado']
        _ajustar_largura_colunas(worksheet_estados)
        _formatar_negrito_linha(worksheet_estados, 2)
        _formatar_colunas_numero(worksheet_estados, 'Volume em R$/mil', 'Quantidade', 'quantidade por contas')

        worksheet_fundos = writer.sheets['Valores por Fundos']
        _ajustar_largura_colunas(worksheet_fundos)
        _formatar_negrito_linha(worksheet_fundos, 2)
        _formatar_colunas_numero(worksheet_fundos, 'Volume R$/mil', 'Quantidade CPFs únicos', 'Quantidade por Contas')

    buffer.seek(0)

    return {
        'ok': True,
        'buffer': buffer,
        'nome_arquivo': nome_arquivo,
        'total_registros': int(len(df_filtrado)),
    }
